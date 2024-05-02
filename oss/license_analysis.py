import os
import pandas as pd
import openpyxl
import json
from tqdm import tqdm
from ordered_set import OrderedSet
from black_duck_api import get_project_components

# name of the Excel file where save the result
WORKSPACE_OCC_DIR = os.path.join(os.getcwd(), 'oss')
FILE_NAME = 'BlackDuckLicenses.xlsx'
RULES_DATA_FRAME = pd.read_excel(
    os.path.join(WORKSPACE_OCC_DIR, 'Rules v073.xlsx'), sheet_name='Licenses')


def save_into_excel(sheet_name: str, set_of_licenses: dict):
    file_path = os.path.join(WORKSPACE_OCC_DIR, FILE_NAME)
    if not os.path.exists(file_path):
        workbook = openpyxl.Workbook()
    else:
        workbook = openpyxl.load_workbook(file_path)
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.create_sheet(sheet_name)

    sheet.cell(row=1, column=1, value="LICENSE")
    sheet.cell(row=1, column=2, value="RULE SET")
    sheet.cell(row=1, column=3, value="NOTES")
    sheet.cell(row=1, column=4, value="BRANCH FOR APPROVAL")
    sheet.cell(row=1, column=5, value="OBBLIGATIONS")
    sheet.cell(row=1, column=6, value="MITIGATION")
    for row, (key, value) in enumerate(set_of_licenses.items(), start=2):
        sheet.cell(row=row, column=1, value=key)
        sheet.cell(row=row, column=2, value=value["rule_set"])

    workbook.save(file_path)
    workbook.close()


def get_rule_set(splitter="Z"*10, license_name="") -> str:
    licenses_splitted = license_name.split(splitter)

    rule_set = []
    for license in licenses_splitted:
        try:
            rule_set.append(
                RULES_DATA_FRAME[RULES_DATA_FRAME["License"] == license]
                ['Rule set'].values[0])
        except IndexError:
            # in case license is not found
            rule_set.append("NOT_FOUND")
    return splitter.join(rule_set)

def simplify_rule_sets(set_of_licenses) -> dict:
    for license in set_of_licenses:
        # Case 1: At least one 'AllApprove' with 'OR' splitter
        if "AllApprove" in set_of_licenses[license]["rule_set"] and " OR " in set_of_licenses[license]["rule_set"]:
            set_of_licenses[license]["rule_set"] = "AllApprove"

        # Case 2: Splitter is 'AND' and 'AllApprove' is present
        set_of_licenses[license]["rule_set"] = set_of_licenses[license]["rule_set"].replace(" AND AllApprove", "")
        set_of_licenses[license]["rule_set"] = set_of_licenses[license]["rule_set"].replace("AllApprove AND ", "")

    return set_of_licenses

def analyze_licenses(module_name: str, module_info: dict, headers: dict) -> dict:
    resp = get_project_components(
        api_headers=headers, project_id=module_info['id'], version_id=module_info['versionId'])
    if 'errorMessage' in resp:
        print(f"Error on get components for {module_name}: {resp['errorMessage']}")
        return None
    else:
        set_of_licenses = {}
        for component in tqdm(
                resp["items"],
                desc=module_name.ljust(15),
                bar_format="{l_bar}{bar:50}{r_bar}{bar:-50b}", ascii=True):
            license_obj = component["licenses"][0]
            license_name = license_obj["licenseDisplay"]
            if license_name[0] == "(":
                # remove parentheses
                license_name = license_name[1:-1]

            if not license_name in set_of_licenses:
                # initialize the dict
                set_of_licenses[license_name] = {
                    "spdx_id": OrderedSet(),
                    "rule_set": "",
                    "components": []
                }
            componentName = f'{component["componentName"]}_{component["componentVersionName"]}'
            license = set_of_licenses[license_name]
            license["components"].append(componentName)

            if len(license_obj["licenses"]) == 0:
                # the component has only 1 licence
                if "spdxId" in license_obj:
                    license["spdx_id"].add(license_obj["spdxId"])
            else:
                # the component has more than 1 licence
                for lic in license_obj["licenses"]:
                    license["spdx_id"].add(lic.get("spdxId", ""))

        # Read data from excel and fill key "rule_set"
        for license_name, license in set_of_licenses.items():
            if " AND " in license_name:  # 2 licenses in AND
                rule_set = get_rule_set(
                    splitter=" AND ", license_name=license_name)
            elif " OR " in license_name:  # 2 licenses in OR
                rule_set = get_rule_set(
                    splitter=" OR ", license_name=license_name)
            else:
                rule_set = get_rule_set(license_name=license_name)
            set_of_licenses[license_name]["rule_set"] = rule_set
            # json can not handle set, so convert to list
            license["spdx_id"] = list(license["spdx_id"])

        # remove useless rulesets
        set_of_licenses = simplify_rule_sets(set_of_licenses)

        # order the licenses
        sorted_licenses = {license: set_of_licenses[license]
                           for license in sorted(set_of_licenses)}

        # save result in excel file
        save_into_excel(sheet_name=module_name,
                        set_of_licenses=sorted_licenses)
        return sorted_licenses


def perform_license_analysis(modules: dict, headers: dict):
    total_set_of_licenses = {}
    for module_name, module_info in modules.items():
        licenses_analyzed = analyze_licenses(
            module_name, module_info, headers)
        if licenses_analyzed:
            total_set_of_licenses[module_name] = licenses_analyzed

    # dict with all licenses regardless the module they belong
    all_licenses = {}
    for module_licenses in total_set_of_licenses.values():
        for license, license_value in module_licenses.items():
            license_copy = license_value.copy()  # by value and not by reference
            license_copy.pop("components")
            all_licenses[license] = license_copy

    # order tot_licenses
    sorted_all_licenses = {
        license: all_licenses[license] for license in sorted(all_licenses)}
    if sorted_all_licenses:
        # save all licenses in excel file
        save_into_excel(sheet_name="All", set_of_licenses=sorted_all_licenses)

        total_set_of_licenses["All"] = sorted_all_licenses
        # save result in json file
        with open(os.path.join(WORKSPACE_OCC_DIR, "licenses_to_check.json"), "w") as outfile:
            json.dump(total_set_of_licenses, outfile, indent=4)
