import os
import pandas as pd
import openpyxl
import json
from tqdm import tqdm
from datetime import datetime
from black_duck_api import get_project_components
from enums import ComponentInteractionRules

# name of the Excel file where save the result
WORKSPACE_OCC_DIR = os.path.join(os.getcwd(), 'oss')
FILE_NAME = 'LitePanelPro-BlackDuckComponents.xlsx'
LICENSES_FILE_PATH = os.path.join(WORKSPACE_OCC_DIR, 'licenses_to_check.json')
RULES_FILE_PATH = os.path.join(WORKSPACE_OCC_DIR, 'Rules v073.xlsx')
RULES_DATA_FRAME = pd.read_excel(RULES_FILE_PATH, sheet_name='Licenses')


def save_licenses_ruleset(sheet_components, licenses_json):
    for row_number_comp, _ in enumerate(sheet_components.iter_rows(min_row=2), start=2):
        for license_name in licenses_json:
            if sheet_components.cell(row=row_number_comp, column=4).value == license_name:
                sheet_components.cell(
                    row=row_number_comp, column=6, value=licenses_json[license_name]["rule_set"])
                sheet_components.cell(
                    row=row_number_comp, column=6, value=licenses_json[license_name]["rule_set"])
                break  # To check if this works because I wrongly used pass before but this should be just fine


def get_decision(workbook, spreadsheet_name, branch_infos,
                 interaction_rule: ComponentInteractionRules):
    """Retrieves a decision based on the Rule set and parameters."""

    if spreadsheet_name == "AllApprove":
        return "approve"

    if spreadsheet_name == "AllReject":
        return "reject"

    if spreadsheet_name in workbook.sheetnames:
        rule_sheet = workbook[spreadsheet_name]
    else:
        return "NOT_FOUND"

    selected_row = 2
    if not branch_infos["Internal"]:
        selected_row = selected_row + 24

    if not branch_infos["Modification"]:
        selected_row = selected_row + 12

    if not branch_infos["Hosted"]:
        selected_row = selected_row + 8
    else:
        if not branch_infos["Host_control_ABB"]:
            selected_row = selected_row + 4

    selected_row = selected_row + interaction_rule.value

    return rule_sheet.cell(row=selected_row, column=6).value


def save_into_excel(sheet_name: str, set_of_components: dict, branch_infos: dict):
    file_path = os.path.join(WORKSPACE_OCC_DIR, FILE_NAME)
    if not os.path.exists(file_path):
        workbook = openpyxl.Workbook()
    else:
        workbook = openpyxl.load_workbook(file_path)
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.create_sheet(sheet_name)

    sheet.cell(row=1, column=1, value="COMPONENT")
    sheet.cell(row=1, column=2, value="COMPONENT VERSION")
    sheet.cell(row=1, column=3, value="USAGE")
    sheet.cell(row=1, column=4, value="LICENSE")
    sheet.cell(row=1, column=5, value="LICENSE RISK")

    for row, (_, value) in enumerate(set_of_components.items(), start=2):
        sheet.cell(row=row, column=1, value=value["COMPONENT"])
        sheet.cell(row=row, column=2, value=value["COMPONENT_VERSION"])
        sheet.cell(row=row, column=3, value=value["USAGE"])
        sheet.cell(row=row, column=4, value=value["LICENSE"])
        sheet.cell(row=row, column=5, value=value["LICENSE_RISK"])

    if not os.path.exists(LICENSES_FILE_PATH):
        sheet.cell(
            row=1, column=6, value="To load column \"RULE SET\", first analyze the licenses of this project.")
    else:
        with open(LICENSES_FILE_PATH) as json_file:
            licenses_json = json.load(json_file)
            # easier to use "All" instead of sheet_name but could be different with more sheets
            licenses_json = licenses_json["All"]

            creation_timestamp = os.path.getctime(LICENSES_FILE_PATH)
            creation_datetime = datetime.fromtimestamp(creation_timestamp)
            header = f"RULESET (last update: {
                creation_datetime.strftime('%Y-%m-%d %H:%M:%S')})"

            sheet.cell(row=1, column=6, value=header)
            save_licenses_ruleset(sheet_components=sheet,
                                  licenses_json=licenses_json)

        if not os.path.exists(RULES_FILE_PATH):
            sheet.cell(
                row=1, column=7, value="To load column \"BRANCH FOR APPROVAL\", add the Rules file inside the folder oss")
        else:
            sheet.cell(row=1, column=7, value="APPROVED ❌✔️🤷‍♀️")
            workbook_rules = openpyxl.load_workbook(RULES_FILE_PATH)
            for row_number_comp, _ in enumerate(sheet.iter_rows(min_row=2),
                                                start=2):
                usage_str = sheet.cell(row=row_number_comp, column=3).value
                # Using column RULESET to calculate decision
                decision = get_decision(workbook_rules,
                                        sheet.cell(row=row_number_comp,
                                                   column=6).value,
                                        branch_infos,
                                        ComponentInteractionRules[usage_str])

                if decision == "reject":
                    sheet.cell(row=row_number_comp, column=7, value="❌")
                    print("component: ",
                          sheet.cell(row=row_number_comp, column=1).value,
                          " doesn't comply.")
                elif decision == "approve":
                    sheet.cell(row=row_number_comp, column=7, value="✔️")
                elif decision == "check":
                    sheet.cell(row=row_number_comp, column=7, value="🤷‍♀️")
                    print("component: ",
                          sheet.cell(row=row_number_comp, column=1).value,
                          " to check.")
                else:
                    sheet.cell(row=row_number_comp,
                               column=7, value="Not found")
                    print("component: ",
                          sheet.cell(row=row_number_comp, column=1).value,
                          " not found.")

    workbook.save(file_path)
    workbook.close()


def get_rule_set(splitter="Z"*10, license_name="") -> str:
    licenses_splitted = license_name.split(splitter)
    rule_set = []
    for license in licenses_splitted:
        try:
            rule_set.append(
                RULES_DATA_FRAME[RULES_DATA_FRAME["License"] == license]
                ['Rule set'].values[0])
        except IndexError:
            # in case license is not found
            rule_set.append("NOT_FOUND")
    return splitter.join(rule_set)


def analyze_components(module_name: str, module_info: dict, headers: dict, branch_infos: dict) -> dict:
    resp = get_project_components(
        api_headers=headers, project_id=module_info['id'], version_id=module_info['versionId'])

    if 'errorMessage' in resp:
        print(f"Error on get components for {
              module_name}: {resp['errorMessage']}")
        return None
    else:
        set_of_components = {}
        for index, component in enumerate(tqdm(
                resp["items"],
                desc=module_name.ljust(15),
                bar_format="{l_bar}{bar:50}{r_bar}{bar:-50b}", ascii=True)):

            set_of_components[index] = {
                "COMPONENT": component["componentName"],
                "COMPONENT_VERSION": component["componentVersionName"],
                "USAGE": "",
                "LICENSE": "",
                "LICENSE_RISK": "",
            }

            if len(component["usages"]) == 1:
                set_of_components[index]["USAGE"] = component["usages"][0]
            else:
                for usg in component["usages"]:
                    set_of_components[index]["USAGE"] += f"{usg} "

            if component["licenses"][0]["licenseDisplay"][0] == "(":
                set_of_components[index]["LICENSE"] = component["licenses"][0]["licenseDisplay"][1:-1]
            else:
                set_of_components[index]["LICENSE"] = component["licenses"][0]["licenseDisplay"]

            for count in component["licenseRiskProfile"]["counts"]:
                if count["countType"] == "UNKNOWN":
                    unknown = count["count"]
                if count["countType"] == "OK":
                    ok = count["count"]
                if count["countType"] == "LOW":
                    low = count["count"]
                if count["countType"] == "MEDIUM":
                    medium = count["count"]
                if count["countType"] == "HIGH":
                    high = count["count"]
                if count["countType"] == "CRITICAL":
                    critical = count["count"]

            risk_counts = {
                "UNKNOWN": unknown,
                "OK": ok,
                "LOW": low,
                "MEDIUM": medium,
                "HIGH": high,
                "CRITICAL": critical
            }

            for risk, count in risk_counts.items():
                if count != 0:
                    set_of_components[index]["LICENSE_RISK"] += f"{risk} "

        # save result in excel file
        save_into_excel(sheet_name=module_name,
                        set_of_components=set_of_components, branch_infos=branch_infos)
        return set_of_components


def perform_components_analysis(modules: dict, headers: dict, branch_infos: dict):
    for module_name, module_info in modules.items():
        analyze_components(module_name, module_info, headers, branch_infos)
